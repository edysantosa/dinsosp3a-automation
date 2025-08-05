from openpyxl import load_workbook
import datetime

# Untuk mengirim pesan ke whatsapp server
import mimetypes
import requests
import base64
import argparse
import os
import sys

if __name__ == "__main__":

    # Load the workbook
    workbook = load_workbook('pegawai.xlsx')

    # Select a specific sheet
    sheet = workbook.worksheets[0]

    # # Iterate through rows and columns to access cell values
    # for row in sheet.iter_rows():
    #     for cell in row:
    #         print(cell.row)

    # # Access a specific cell
    # cell_value = sheet['A1'].value
    # print(f"Value of cell A1: {cell_value}")


    present = datetime.datetime.now()
    birthdays = []

    # Iterate birthday columns
    for column_cells in sheet.iter_cols(min_col=2, max_col=2):
        for cell in column_cells:
            # print(cell.value)
            # print(datetime.datetime.strptime(cell.value, "%d-%m-%Y"))
            # if (present.date() == cell.value.date()):
            if cell.value.date().day == present.date().day and cell.value.date().month == present.date().month:
                # print(sheet.cell(row=cell.row, column=cell.column-1).value)
                birthdays.append({'sort': sheet.cell(row=cell.row, column=cell.column+1).value, 'name': sheet.cell(row=cell.row, column=cell.column-1).value})

    birthdays.sort(key= lambda x : x['sort'])

    # for person in birthdays:
    #     print(person)

    if len(birthdays) == 0:
        print("gak ada yang ultah")
        sys.exit(0)

    message = f"*Selamat ulang tahun untuk*:\n\n{''.join(['- _{}_\n'.format(x['name']) for x in birthdays])}\n*Semoga sehat dan sukses selalu*"
    # print(message)

    # headers = {"Content-Type": "application/json;charset=utf-8"}
    # json_payload = {
    #     "group_name": groupName,
    #     "message": message
    # }

    # url = 'http://localhost:8000/send-group-message'
    # r = requests.post(url, headers=headers, json=json_payload)
    # print (r.json())
    
    try:
        with open("happy_birthday.png", "rb") as pdf_file:
            encoded_string = base64.b64encode(pdf_file.read())

        headers = {"Content-Type": "application/json;charset=utf-8"}
        json_payload = {
            "group_name":"Dinas Sosial P3A Prov. Bali",
            "message":message,
            "file_name": "happy_birthday.png",
            "mime_type": mimetypes.guess_file_type("happy_birthday.png")[0],
            "file": encoded_string.decode("utf-8")
        }
        
        url = 'http://localhost:8000/send-group-message'
        r = requests.post(url, headers=headers, json=json_payload)
        print (r)
    except OSError as e:
        print("File gambar tidak ditemukan")
    except requests.exceptions.RequestException as e:
        # catastrophic error. bail.
        raise SystemExit(e)
